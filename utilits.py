import inspect
import logging
import openpyxl
import time
import re
from functools import wraps


console_out = logging.StreamHandler()
file_log = logging.FileHandler("application.log", mode="w")
logging.basicConfig(handlers=(file_log, console_out), level=logging.INFO,
                    format='[%(asctime)s | %(levelname)s]: %(message)s')


def log_print(func):
    @wraps(func)
    def _wrapper(*args, **kwargs):
        sig = inspect.signature(func)
        bound = sig.bind(*args, **kwargs)
        bound.apply_defaults()
        all_args = {arg: value for arg, value in bound.arguments.items() if arg != 'self'}  # Создаём словарь аргументов, исключая 'self'
        logging.info(f'CALL {func.__name__}{all_args}...')
        start_time = time.time()

        result = func(*args, **kwargs)

        exec_time = time.time() - start_time
        logging.info(f'RESULT {func.__name__}{all_args} ({exec_time:.3f} sec.):\n{result} ')
        return result
    return _wrapper


def find_phone_numbers(text):
    pattern = r'\b\d{5,15}\b'
    return re.findall(pattern, text)


def normalize_phone(number):
    number = re.sub("[(|)|\-| ]", "", number)
    if number.startswith('+7'):
        number = '8' + number[2:]
    return number

@log_print
def read_contacts_xlsx(file_path: str):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Выбор активного листа
    phone_book = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Начинаем со второй строки, если есть заголовок
        client_name = row[0]  # Название или ФИО из первого столбца
        for cell in row[4:8]:  # Номера расположены с 4-й по 8-й столбец
            if cell is not None:
                numbers = cell.split(';')
                for number in numbers:
                    found_numbers = find_phone_numbers(number)  # Очистка и поиск номеров телефонов
                    for found_number in found_numbers:
                        phone_book[found_number] = client_name
    return phone_book


@log_print
def read_employee_xlsx(file_path: str):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    phone_book = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        fio, position, work_phone, mobile_phone = [cell if cell is not None else "" for cell in row[:4]]
        fio = fio.strip()
        position = position.strip()
        if not fio or (not work_phone and not mobile_phone):
            continue

        phone_numbers = []
        if work_phone:
            phone_numbers.extend(work_phone.split(','))
        if mobile_phone:
            phone_numbers.extend(mobile_phone.split(','))

        value_parts = [fio]
        if position:
            value_parts.append(position)
        value = ', '.join(value_parts)

        for phone_number in phone_numbers:
            phone = normalize_phone(phone_number.strip())
            if len(phone) > 6:
                phone_book[phone] = value
    return phone_book


if __name__ == '__main__':
    file_path = '/home/user/Autorun/SBIS_contacts/Клиенты_CRM.xlsx'
    sbis_contacts = read_contacts_xlsx(file_path)
    for phone, name in sbis_contacts.items():
        print(f"{phone}: {name}")
